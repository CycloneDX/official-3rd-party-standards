"""
Python module that reads in the NIST SSDF and will generate
the CycloneDX Attestation for the standard.
"""

from uuid import uuid4

from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook


class Task:  # pylint: disable=too-few-public-methods
    """
    Class containing SSDF "task" metadata
    """

    def __init__(self, sheet, cell, examples_cell, references_cell):
        self.identifier, self.description = cell.value.split(": ", 1)
        self.examples = examples_cell.value
        self.references = references_cell.value
        self.coordinate = f"{sheet.title}!{cell.coordinate}"


class Practice:  # pylint: disable=too-few-public-methods
    """
    Class containing SSDF "practice" metadata
    """

    def __init__(self, sheet, cell):
        title_identifier, self.description = cell.value.split(": ", 1)
        self.title, identifier = title_identifier.split("(")
        self.identifier = identifier[0:-1]
        self.group = self.identifier.split(".", 1)[0]
        self.coordinate = f"{sheet.title}!{cell.coordinate}"
        self.tasks = []


class Group(Practice):  # pylint: disable=too-few-public-methods
    """
    Class containing SSDF "group" metadata
    """

    def __init__(self, sheet, cell):
        super().__init__(sheet, cell)
        self.practices = []


class Standard:  # pylint: disable=too-few-public-methods
    """
    Class containing SSDF "standards" metadata
    """

    def __init__(self):
        self.identifier = "ssdf-1.1"
        self.name = "Secure Software Development Framework (SSDF) Version 1.1"
        self.description = "NIST Special Publication 800-218"
        self.version = "1.1"
        self.owner = "National Institute of Standards and Technology"
        self.url = "https://csrc.nist.gov/csrc/media/Publications/sp/800-218/final/documents/NIST.SP.800-218.SSDF-table.xlsx"  # pylint: disable=line-too-long


def parse_ssdf():
    """
    Function that reads in the SSDF requirement spreadsheet and parses it into
    Practices with Requirements.

    Returns:
        list[Practices]: The list of practices with requirements from the SSDF
                         spreadsheet
    """
    workbook = load_workbook(filename="NIST.SP.800-218.SSDF-table.xlsx")

    worksheet = workbook["Groups"]

    groups = {}
    for row in worksheet.iter_rows(max_col=1):
        group = Group(worksheet, row[0])
        groups[group.identifier] = group

    worksheet = workbook["SSDF"]

    last_practice: Practice = None

    for row in worksheet.iter_rows(min_row=2, max_col=4):
        if row[0].value:
            if last_practice:
                groups[last_practice.group].practices.append(last_practice)
            last_practice = Practice(worksheet, row[0])

        task = Task(worksheet, row[1], row[2], row[3])
        last_practice.tasks.append(task)

    groups[last_practice.group].practices.append(last_practice)
    return groups


def generator():
    """
    Top level function that ...
    - reads in the SSDF excel spreadsheet
    - passes data to a jinja2 template
    """
    standard = Standard()

    # Practices are a SSDF specific term for top level requirements
    groups = parse_ssdf()

    environment = Environment(
        loader=FileSystemLoader("attestation_generator/templates")
    )
    template = environment.get_template("ssdf-1.1.json.template")
    content = template.render(
        serialNumber=uuid4().urn,
        standard=standard,
        groups=groups.values(),
        autoescape=True,
    )
    with open("ssdf-1.1.cdx.json", mode="w", encoding="utf-8") as message:
        message.write(content)


if __name__ == "__main__":  # pragma: no cover
    generator()
